"""
VFS 资产管理 — 三表统一数据录入工作流
============================================
支持向以下三个表添加数据：
  1. Asset ID Status_FAR.xlsx  (按城市分sheet)
  2. China Spare Asset.xlsx      (Raw Data sheet)
  3. VFS IT Inventory Guangzhou 2026.xlsx (按部门分sheet)

使用方式：
  from asset_workflow import AssetBuilder, add_assets, AssetRecord

  record = (AssetBuilder()
      .city('GZ').office('VFS').department('Canada')
      .equipment('Desktop').brand('HP').model('288 G9').serial_no('SN-001')
      .status('In Use').user_name('Tony')
      .build())

  add_assets([record])
"""

import shutil
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ============================================================
# 路径配置 — 修改此处指向实际文件位置
# ============================================================
DEFAULT_FAR_PATH   = r"C:\Users\Administrator\Desktop\py\Asset ID Status_FAR.xlsx"
DEFAULT_SPARE_PATH = r"C:\Users\Administrator\Desktop\py\China Spare Asset.xlsx"
DEFAULT_INV_PATH   = r"C:\Users\Administrator\Desktop\py\VFS IT Inventory Guangzhou 2026.xlsx"

# 工作目录输出(可选)
OUTPUT_DIR = r"C:\Users\Administrator\Desktop\py\output"

# ============================================================
# 统一数据模型
# ============================================================
class AssetRecord:
    """统一的资产数据模型，涵盖三个表的所有字段。"""
    def __init__(self, **kwargs):
        # 公共字段
        self.city = kwargs.get('city', '')
        self.office = kwargs.get('office', '')
        self.department = kwargs.get('department', '')
        self.type_of_equipment = kwargs.get('type_of_equipment', '')
        self.brand = kwargs.get('brand', '')
        self.model = kwargs.get('model', '')
        self.serial_no = kwargs.get('serial_no', '')
        self.date_of_put_into_use = kwargs.get('date_of_put_into_use', '')
        self.cost_center = kwargs.get('cost_center', '')
        self.status_of_machine = kwargs.get('status_of_machine', '')
        self.asset_id_status = kwargs.get('asset_id_status', '')
        self.asset_id = kwargs.get('asset_id', '')
        self.cc_id = kwargs.get('cc_id', '')
        self.entity = kwargs.get('entity', '')
        self.entity_name = kwargs.get('entity_name', '')
        self.global_asset_tag = kwargs.get('global_asset_tag', '')
        self.user_name = kwargs.get('user_name', '')
        self.hostname = kwargs.get('hostname', '')
        self.ip_address = kwargs.get('ip_address', '')
        self.warranty_amc = kwargs.get('warranty_amc', '')
        self.expiry_date = kwargs.get('expiry_date', '')
        self.processor = kwargs.get('processor', '')
        self.speed = kwargs.get('speed', '')
        self.ram = kwargs.get('ram', '')
        self.hdd = kwargs.get('hdd', '')
        self.size = kwargs.get('size', '')
        self.os = kwargs.get('os', '')
        self.remark = kwargs.get('remark', '')

        # 表2 额外字段
        self.rm_name = kwargs.get('rm_name', '')
        self.owner = kwargs.get('owner', '')
        self.location = kwargs.get('location', '')
        self.city2 = kwargs.get('city2', '')
        self.office2 = kwargs.get('office2', '')
        self.dept2 = kwargs.get('dept2', '')
        self.recipient = kwargs.get('recipient', '')
        self.sr_no = kwargs.get('sr_no', None)

        # 表3 额外字段
        self.fmc_or_not = kwargs.get('fmc_or_not', '')
        self.purchase_date = kwargs.get('purchase_date', '')
        self.vendor = kwargs.get('vendor', '')
        self.purchase_value = kwargs.get('purchase_value', '')
        self.purchase_entity = kwargs.get('purchase_entity', '')
        self.purchase_value_ref = kwargs.get('purchase_value_ref', '')

    def to_dict(self):
        return {k: v for k, v in vars(self).items() if v is not None and v != ''}

    def __repr__(self):
        return f"AssetRecord({self.type_of_equipment} | {self.brand} {self.model} | SN:{self.serial_no})"


# ============================================================
# 链式构建器
# ============================================================
class AssetBuilder:
    """链式构建器，简化 AssetRecord 的创建。"""
    def __init__(self):
        self._data = {}

    def city(self, v):        self._data['city'] = v; return self
    def office(self, v):      self._data['office'] = v; return self
    def department(self, v):   self._data['department'] = v; return self
    def equipment(self, v):   self._data['type_of_equipment'] = v; return self
    def type_of_equipment(self, v): self._data['type_of_equipment'] = v; return self
    def brand(self, v):       self._data['brand'] = v; return self
    def model(self, v):       self._data['model'] = v; return self
    def serial_no(self, v):   self._data['serial_no'] = v; return self
    def date_of_put_into_use(self, v): self._data['date_of_put_into_use'] = v; return self
    def cost_center(self, v): self._data['cost_center'] = v; return self
    def status(self, v):      self._data['status_of_machine'] = v; return self
    def status_of_machine(self, v): self._data['status_of_machine'] = v; return self
    def asset_id(self, v):    self._data['asset_id'] = v; return self
    def asset_id_status(self, v): self._data['asset_id_status'] = v; return self
    def cc_id(self, v):       self._data['cc_id'] = v; return self
    def entity(self, v):      self._data['entity'] = v; return self
    def entity_name(self, v): self._data['entity_name'] = v; return self
    def global_asset_tag(self, v): self._data['global_asset_tag'] = v; return self
    def user_name(self, v):   self._data['user_name'] = v; return self
    def hostname(self, v):    self._data['hostname'] = v; return self
    def ip_address(self, v):  self._data['ip_address'] = v; return self
    def warranty_amc(self, v): self._data['warranty_amc'] = v; return self
    def warranty(self, v):    self._data['warranty_amc'] = v; return self
    def expiry_date(self, v): self._data['expiry_date'] = v; return self
    def processor(self, v):   self._data['processor'] = v; return self
    def speed(self, v):       self._data['speed'] = v; return self
    def ram(self, v):         self._data['ram'] = v; return self
    def hdd(self, v):         self._data['hdd'] = v; return self
    def size(self, v):        self._data['size'] = v; return self
    def os(self, v):          self._data['os'] = v; return self
    def remark(self, v):      self._data['remark'] = v; return self

    # 表2
    def rm_name(self, v):     self._data['rm_name'] = v; return self
    def owner(self, v):       self._data['owner'] = v; return self
    def location(self, v):    self._data['location'] = v; return self
    def city2(self, v):       self._data['city2'] = v; return self
    def office2(self, v):     self._data['office2'] = v; return self
    def dept2(self, v):       self._data['dept2'] = v; return self
    def recipient(self, v):   self._data['recipient'] = v; return self

    # 表3
    def fmc_or_not(self, v):  self._data['fmc_or_not'] = v; return self
    def purchase_date(self, v): self._data['purchase_date'] = v; return self
    def vendor(self, v):       self._data['vendor'] = v; return self
    def purchase_value(self, v): self._data['purchase_value'] = v; return self
    def purchase_entity(self, v): self._data['purchase_entity'] = v; return self
    def purchase_value_ref(self, v): self._data['purchase_value_ref'] = v; return self

    def build(self) -> AssetRecord:
        return AssetRecord(**self._data)


# ============================================================
# 表1: Asset ID Status_FAR.xlsx
# ============================================================
FAR_COL_MAP = {
    'city': 'A', 'office': 'B', 'department': 'C', 'type_of_equipment': 'D',
    'brand': 'E', 'model': 'F', 'serial_no': 'G', 'date_of_put_into_use': 'H',
    'cost_center': 'I', 'status_of_machine': 'J', 'asset_id_status': 'K',
    'asset_id': 'L', 'cc_id': 'M', 'entity': 'N', 'entity_name': 'O',
    'global_asset_tag': 'P', 'user_name': 'Q', 'hostname': 'R',
    'ip_address': 'S', 'warranty_amc': 'T', 'expiry_date': 'U',
    'processor': 'V', 'speed': 'W', 'ram': 'X', 'hdd': 'Y', 'size': 'Z',
    'os': 'AA', 'remark': 'AB'
}
FAR_HEADER_ROWS = 3


def write_to_far(record: AssetRecord, filepath: str, sheet_name: str = None) -> bool:
    target_sheet = sheet_name or record.city
    wb = load_workbook(filepath)
    if target_sheet not in wb.sheetnames:
        print(f"  ⚠️  表1: sheet '{target_sheet}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[target_sheet]
    next_row = ws.max_row + 1

    for field, col_letter in FAR_COL_MAP.items():
        val = getattr(record, field, '')
        if val is not None and val != '':
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表1 [{target_sheet}] 行 {next_row}: 写入成功")
    return True


# ============================================================
# 表2: China Spare Asset.xlsx
# ============================================================
SPARE_COL_MAP = {
    'sr_no': 'A', 'rm_name': 'B', 'owner': 'C', 'location': 'D',
    'city': 'E', 'office': 'F', 'department': 'G', 'type_of_equipment': 'H',
    'user_name': 'I', 'hostname': 'J', 'brand': 'K', 'model': 'L',
    'serial_no': 'M', 'date_of_put_into_use': 'N', 'city2': 'O',
    'office2': 'P', 'dept2': 'Q', 'status_of_machine': 'R',
    'recipient': 'S', 'remark': 'T'
}


def write_to_spare(record: AssetRecord, filepath: str, auto_sr_no: int = None) -> bool:
    wb = load_workbook(filepath)
    ws = wb['Raw Data']
    next_row = ws.max_row + 1

    if auto_sr_no is not None:
        record.sr_no = auto_sr_no
    elif record.sr_no is None:
        max_no = 0
        for r in range(2, next_row):
            v = ws.cell(r, 1).value
            if v and str(v).isdigit():
                max_no = max(max_no, int(v))
        record.sr_no = max_no + 1

    for field, col_letter in SPARE_COL_MAP.items():
        val = getattr(record, field, '')
        if val is not None and val != '':
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表2 [Raw Data] 行 {next_row}: 写入成功 (Sr.No={record.sr_no})")
    return True


# ============================================================
# 表3: VFS IT Inventory Guangzhou 2026.xlsx
# ============================================================
INV_COL_MAP = {
    'sr_no': 'A', 'rm_name': 'B', 'fmc_or_not': 'C', 'location': 'D',
    'city': 'E', 'office': 'F', 'department': 'G', 'type_of_equipment': 'H',
    'user_name': 'I', 'hostname': 'J', 'ip_address': 'K', 'brand': 'L',
    'model': 'M', 'serial_no': 'N', 'asset_id': 'O', 'global_asset_tag': 'P',
    'purchase_date': 'Q', 'warranty_amc': 'R', 'expiry_date': 'S',
    'vendor': 'T', 'status_of_machine': 'U', 'processor': 'V', 'speed': 'W',
    'ram': 'X', 'hdd': 'Y', 'size': 'Z', 'os': 'AA', 'remark': 'AB',
    'purchase_value': 'AC', 'purchase_entity': 'AD', 'purchase_value_ref': 'AE'
}

# 部门 → sheet 映射
DEPT_SHEET_MAP = {
    'canada':       ['Canada'],
    'germany':      ['Ger&JVAC&Swiss'],
    'swiss':        ['Ger&JVAC&Swiss'],
    'jvac':         ['Ger&JVAC&Swiss'],
    'italy':        ['Italy'],
    'non-sch':      ['Non-Sch'],
    'platinum':     ['Platinum Lounge'],
    'ro':           ['RO'],
    'uk':           ['UK&Ireland'],
    'ireland':      ['UK&Ireland'],
    'us':           ['US'],
    'vasco':        ['Vasco'],
}


def infer_sheet_name(record: AssetRecord) -> str:
    """根据部门自动推断目标 sheet 名。"""
    dept_lower = record.department.lower()
    for key, sheets in DEPT_SHEET_MAP.items():
        if key in dept_lower:
            return sheets[0]
    return 'Canada'  # 默认


def write_to_inventory(record: AssetRecord, filepath: str,
                      sheet_name: str = None, auto_sr_no: int = None) -> bool:
    wb = load_workbook(filepath)

    if sheet_name is None:
        sheet_name = infer_sheet_name(record)

    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  表3: sheet '{sheet_name}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    if auto_sr_no is not None:
        record.sr_no = auto_sr_no
    else:
        max_no = 0
        for r in range(2, next_row):
            v = ws.cell(r, 1).value
            if v and str(v).isdigit():
                max_no = max(max_no, int(v))
        record.sr_no = max_no + 1

    for field, col_letter in INV_COL_MAP.items():
        val = getattr(record, field, '')
        if val is not None and val != '':
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表3 [{sheet_name}] 行 {next_row}: 写入成功 (Sr.No={record.sr_no})")
    return True


# ============================================================
# 统一入口
# ============================================================
def add_assets(
    records: list,
    far_path: str = DEFAULT_FAR_PATH,
    spare_path: str = DEFAULT_SPARE_PATH,
    inv_path: str = DEFAULT_INV_PATH,
    target_tables: list = None,
    far_sheet: str = None,
    inv_sheet: str = None,
    copy_to_output: bool = False
) -> dict:
    """
    统一入口：向指定的表添加资产记录。

    Args:
        records: AssetRecord 列表
        target_tables: 要写入的表, ['far','spare','inv'] 的子集, None=全部
        far_sheet: 表1的目标sheet(城市), None则使用 record.city
        inv_sheet: 表3的目标sheet(部门), None则自动推断
        copy_to_output: 是否将结果复制到 /data/workspace/output/

    Returns:
        dict: {'far': [...], 'spare': [...], 'inv': [...]}
    """
    if target_tables is None:
        target_tables = ['far', 'spare', 'inv']

    if copy_to_output:
        import os
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        shutil.copy2(far_path, os.path.join(OUTPUT_DIR, "Asset ID Status_FAR.xlsx"))
        shutil.copy2(spare_path, os.path.join(OUTPUT_DIR, "China Spare Asset.xlsx"))
        shutil.copy2(inv_path, os.path.join(OUTPUT_DIR, "VFS IT Inventory Guangzhou 2026.xlsx"))
        far_path = os.path.join(OUTPUT_DIR, "Asset ID Status_FAR.xlsx")
        spare_path = os.path.join(OUTPUT_DIR, "China Spare Asset.xlsx")
        inv_path = os.path.join(OUTPUT_DIR, "VFS IT Inventory Guangzhou 2026.xlsx")

    print(f"\n{'='*65}")
    print(f"  VFS 资产数据录入工作流")
    print(f"  共 {len(records)} 条记录 → 目标: {target_tables}")
    print(f"{'='*65}")

    results = {'far': [], 'spare': [], 'inv': []}

    for i, rec in enumerate(records):
        print(f"\n--- 记录 #{i+1} ---")
        print(f"  {rec}")

        if 'far' in target_tables:
            print("  >> 表1 (Asset ID Status_FAR)...")
            ok = write_to_far(rec, far_path, far_sheet)
            results['far'].append(ok)

        if 'spare' in target_tables:
            print("  >> 表2 (China Spare Asset)...")
            ok = write_to_spare(rec, spare_path)
            results['spare'].append(ok)

        if 'inv' in target_tables:
            print("  >> 表3 (VFS IT Inventory)...")
            ok = write_to_inventory(rec, inv_path, inv_sheet)
            results['inv'].append(ok)

    # 汇总
    print(f"\n{'='*65}")
    print("  写入汇总:")
    for tbl, res in results.items():
        if res:
            success = sum(1 for r in res if r)
            print(f"   {tbl}: {success}/{len(res)} 成功")
    print(f"{'='*65}\n")

    return results


# ============================================================
# 批量导入：从字典列表构建记录
# ============================================================
def from_dict_list(dict_list: list) -> list:
    """从字典列表批量构建 AssetRecord 列表。"""
    return [AssetRecord(**d) for d in dict_list]


# ============================================================
# 快捷方式：仅写入指定表
# ============================================================
def add_to_far_only(records: list, sheet_name: str = None,
                    far_path: str = DEFAULT_FAR_PATH, copy_to_output: bool = False):
    """仅写入表1 (Asset ID Status_FAR)。"""
    return add_assets(records, far_path=far_path, target_tables=['far'],
                      far_sheet=sheet_name, copy_to_output=copy_to_output)

def add_to_spare_only(records: list, spare_path: str = DEFAULT_SPARE_PATH,
                      copy_to_output: bool = False):
    """仅写入表2 (China Spare Asset)。"""
    return add_assets(records, spare_path=spare_path, target_tables=['spare'],
                      copy_to_output=copy_to_output)

def add_to_inv_only(records: list, sheet_name: str = None,
                    inv_path: str = DEFAULT_INV_PATH, copy_to_output: bool = False):
    """仅写入表3 (VFS IT Inventory)。"""
    return add_assets(records, inv_path=inv_path, target_tables=['inv'],
                      inv_sheet=sheet_name, copy_to_output=copy_to_output)


# ============================================================
# 演示 & 自测
# ============================================================
if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════════════╗
║       VFS 资产管理 — 三表统一数据录入工作流 (已就绪)            ║
╚══════════════════════════════════════════════════════════════╝
    """)

    # 复制文件到工作目录
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    shutil.copy2(DEFAULT_FAR_PATH, os.path.join(OUTPUT_DIR, "Asset ID Status_FAR.xlsx"))
    shutil.copy2(DEFAULT_SPARE_PATH, os.path.join(OUTPUT_DIR, "China Spare Asset.xlsx"))
    shutil.copy2(DEFAULT_INV_PATH, os.path.join(OUTPUT_DIR, "VFS IT Inventory Guangzhou 2026.xlsx"))

    # 构建多条演示记录
    records = [
        (AssetBuilder()
            .city('GZ').office('VFS').department('Canada')
            .equipment('Desktop').brand('HP').model('280 G9 Demo').serial_no('DEMO-001')
            .date_of_put_into_use('2026-6')
            .status('In Use').user_name('Demo User 1')
            .processor('Intel Core i7').speed('3.0GHz').ram('16G')
            .hdd('SSD').size('512G').os('WIN11')
            .remark('演示数据')
            .rm_name('Tony').owner('VFS').location('Guangzhou')
            .fmc_or_not('VFS').purchase_date('2026-6-28')
            .build()),

        (AssetBuilder()
            .city('BJ').office('VFS').department('UK')
            .equipment('Laptop').brand('Dell').model('Latitude 7440').serial_no('DEMO-002')
            .date_of_put_into_use('2026-6')
            .status('Spare').user_name('Demo User 2')
            .rm_name('Tony').owner('VFS').location('Beijing')
            .fmc_or_not('VFS')
            .build()),
    ]

    # 执行写入（写入到 output 目录）
    results = add_assets(
        records,
        target_tables=['far', 'spare', 'inv'],
        far_sheet=None,       # 自动使用 record.city
        inv_sheet=None,       # 自动推断
        copy_to_output=True
    )

    print("✅ 工作流执行完成!")
    print(f"   输出文件在 {OUTPUT_DIR} 目录下")
    print("\n💡 使用提示:")
    print("   - 修改 DEFAULT_FAR_PATH 等路径变量指向你的实际文件")
    print("   - 使用 AssetBuilder 链式构建记录（推荐）")
    print("   - 使用 add_assets() 统一写入，或 add_to_xxx_only() 单独写入")
    print("   - target_tables 参数控制写入哪些表")
    print("   - far_sheet / inv_sheet 可手动指定目标sheet，或自动推断")
