"""
导入 photoexe.xlsx 数据到表1和表3
=====================================
使用方式：
1. 修改下面的 DEFAULT_VALUES（补充必填字段）
2. 运行脚本：python import_photoexe.py
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import shutil
import os
from openpyxl.styles import Font, Alignment, Border, Side
import logging
from datetime import datetime

# ============================================================
# 样式定义
# ============================================================
def apply_style(cell, font_size=8):
    """给单元格应用样式：Calibri, 指定大小, 黑色边框"""
    # 字体：Calibri, 指定大小
    cell.font = Font(name='Calibri', size=font_size)

    # 边框：黑色细边框
    thin = Side(style='thin', color='000000')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 垂直居中
    cell.alignment = Alignment(vertical='center')


# ============================================================
# 文件路径
# ============================================================
PHOTOEXE_PATH = r"C:\Users\Administrator\Desktop\py\photoexe.xlsx"
FAR_PATH   = r"C:\Users\Administrator\Desktop\py\Asset ID Status_FAR.xlsx"
INV_PATH   = r"C:\Users\Administrator\Desktop\py\VFS IT Inventory Guangzhou 2026.xlsx"
BACKUP_DIR = r"C:\Users\Administrator\Desktop\py\backup"

# ============================================================
# 日志配置
# ============================================================
LOG_FILE = r"C:\Users\Administrator\Desktop\txcode\asset_workflow—1782576943454\edit\asset_receive.txt"

# 配置日志记录
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 备份功能
# ============================================================
def backup_files():
    """执行前备份原文件，加上时间戳"""
    import shutil
    from datetime import datetime
    
    # 创建备份目录
    os.makedirs(BACKUP_DIR, exist_ok=True)
    
    # 生成时间戳
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 备份文件
    backed_up = []
    for filepath in [FAR_PATH, INV_PATH]:
        if os.path.exists(filepath):
            filename = os.path.basename(filepath)
            name, ext = os.path.splitext(filename)
            backup_name = f"{name}_{timestamp}{ext}"
            backup_path = os.path.join(BACKUP_DIR, backup_name)
            shutil.copy2(filepath, backup_path)
            backed_up.append(backup_path)
            logger.info(f"已备份: {filepath} -> {backup_path}")
    
    print(f"[备份] 原文件已备份到: {BACKUP_DIR}")
    logger.info(f"备份完成，共备份 {len(backed_up)} 个文件")
    return backed_up

# ============================================================
# 配置 - 根据数据自动判断
# ============================================================

# vac 到部门的映射（vac列的值 → 表3的sheet名）
VAC_TO_DEPT = {
    'UK': 'UK&Ireland',
    'US': 'US',
    'Canada': 'Canada',
    'Germany': 'Ger&JVAC&Swiss',
    'Italy': 'Italy',
    'Vasco': 'Vasco',
    'RO': 'RO',
    'Platinum': 'Platinum Lounge',
    'Non-Sch': 'Non-Sch',
}

# model 到设备类型的映射（可继续补充）
MODEL_TO_EQUIPMENT = {
    '324PV': 'Monitor',
    '324pv': 'Monitor',
    '324Pv': 'Monitor',
    # 可继续添加：'ThinkPad': 'Laptop', etc.
}

# 默认值
DEFAULT_VALUES = {
    'office': 'VFS',                 # 办公室
    'status_of_machine': 'In Use',    # 状态
    'user_name': '',                  # 使用人
    'rm_name': '',                    # 房间名
    'location': 'Guangzhou',         # 位置（可根据city自动调整）
    'fmc_or_not': 'VFS',             # FMC
    'purchase_date': '2026-6',        # 购买日期
}

# ============================================================
# 列映射
# ============================================================
FAR_COL_MAP = {
    'city': 'A', 'office': 'B', 'department': 'C', 'type_of_equipment': 'D',
    'brand': 'E', 'model': 'F', 'serial_no': 'G', 'date_of_put_into_use': 'H',
    'cost_center': 'I', 'status_of_machine': 'J', 'asset_id_status': 'K',
    'asset_id': 'L', 'cc_id': 'M', 'entity': 'N', 'entity_name': 'O',
    'global_asset_tag': 'P', 'user_name': 'Q', 'hostname': 'R',
    'ip_address': 'S', 'warranty_amc': 'T', 'expiry_date': 'U',
    'processor': 'V', 'speed': 'W', 'ram': 'X', 'hdd': 'Y', 'size': 'Z',
    'os': 'AA', 'remark': 'AB'
}

INV_COL_MAP = {
    'sr_no': 'A', 'rm_name': 'B', 'fmc_or_not': 'C', 'location': 'D',
    'city': 'E', 'office': 'F', 'department': 'G', 'type_of_equipment': 'H',
    'user_name': 'I', 'hostname': 'J', 'ip_address': 'K', 'brand': 'L',
    'model': 'M', 'serial_no': 'N', 'asset_id': 'O', 'global_asset_tag': 'P',
    'purchase_date': 'Q', 'warranty_amc': 'R', 'expiry_date': 'S',
    'vendor': 'T', 'status_of_machine': 'U', 'processor': 'V', 'speed': 'W',
    'ram': 'X', 'hdd': 'Y', 'size': 'Z', 'os': 'AA', 'remark': 'AB',
    'purchase_value': 'AC', 'purchase_entity': 'AD', 'purchase_value_ref': 'AE'
}

DEPT_SHEET_MAP = {
    'canada': 'Canada',
    'germany': 'Ger&JVAC&Swiss',
    'swiss': 'Ger&JVAC&Swiss',
    'jvac': 'Ger&JVAC&Swiss',
    'italy': 'Italy',
    'non-sch': 'Non-Sch',
    'platinum': 'Platinum Lounge',
    'ro': 'RO',
    'uk': 'UK&Ireland',
    'ireland': 'UK&Ireland',
    'us': 'US',
    'vasco': 'Vasco',
}


def infer_equipment_type(model):
    """根据 model 推断设备类型"""
    if not model:
        return 'Desktop'  # 默认
    model_str = str(model).strip()
    # 精确匹配
    if model_str in MODEL_TO_EQUIPMENT:
        return MODEL_TO_EQUIPMENT[model_str]
    # 不区分大小写匹配
    for key, value in MODEL_TO_EQUIPMENT.items():
        if model_str.lower() == key.lower():
            return value
    # 默认
    return 'Desktop'


def infer_department(vac):
    """根据 vac 推断部门（表3的sheet名）"""
    # 错误判断：vac 为空值则报错
    if not vac or (isinstance(vac, str) and vac.strip() == ''):
        raise ValueError("错误: vac 不能为空值！请在 photoexe.xlsx 中检查 B 列数据。")
    
    vac_str = str(vac).strip()
    if vac_str in VAC_TO_DEPT:
        result = VAC_TO_DEPT[vac_str]
        # 如果是 UK&Ireland，返回 UK
        if result == 'UK&Ireland':
            return 'UK'
        return result
    # 不区分大小写匹配
    for key, value in VAC_TO_DEPT.items():
        if vac_str.lower() == key.lower():
            # 如果是 UK&Ireland，返回 UK
            if value == 'UK&Ireland':
                return 'UK'
            return value
    return 'UK'  # 默认改为 UK


def read_photoexe():
    """读取 photoexe.xlsx 的 Asset_Receive sheet，自动推断字段值"""
    wb = load_workbook(PHOTOEXE_PATH)
    ws = wb['Asset_Receive']  # 读取 Asset_Receive sheet

    data_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # 跳过空行
            continue

        city = row[0] if len(row) > 0 else 'GZ'        # A列: city
        vac = row[1] if len(row) > 1 else ''             # B列: vac
        sn = row[2] if len(row) > 2 else ''              # C列: sn
        brand = row[3] if len(row) > 3 else ''           # D列: brand
        model = row[4] if len(row) > 4 else ''           # E列: model
        globaltag = row[5] if len(row) > 5 else ''       # F列: globaltag
        assetid = row[6] if len(row) > 6 else ''         # G列: assetid

        # 自动推断
        department = infer_department(vac)
        equipment_type = infer_equipment_type(model)

        # 构建数据字典
        mapped = {
            'city': city,
            'department': department,
            'type_of_equipment': equipment_type,
            'serial_no': sn,
            'brand': brand,
            'model': model,
            'global_asset_tag': globaltag,
            'asset_id': assetid,
        }

        # 合并默认值
        for k, v in DEFAULT_VALUES.items():
            if v:  # 只添加非空默认值
                mapped[k] = v

        data_list.append(mapped)

    wb.close()
    return data_list


def write_to_far(data, filepath=FAR_PATH):
    """写入表1"""
    sheet_name = data.get('city', 'GZ')
    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        print(f"  [WARNING]  表1: sheet '{sheet_name}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    for field, col_letter in FAR_COL_MAP.items():
        col_idx = column_index_from_string(col_letter)
        val = data.get(field, '')
        if not val:
            val = 'NA'
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        apply_style(cell, font_size=8)  # 表1: Calibri 8号

    wb.save(filepath)
    wb.close()
    print(f"  [OK] 表1 [{sheet_name}] 行 {next_row}: 写入成功")
    logger.info(f"    表1: 在 sheet '{sheet_name}' 第 {next_row} 行写入数据")
    return True


def write_to_inventory(data, filepath=INV_PATH):
    """写入表3"""
    wb = load_workbook(filepath)

    # 推断 sheet
    dept = data.get('department', '')
    sheet_name = ''

    # 如果是 UK，则查找 UK&Ireland sheet
    if dept == 'UK':
        sheet_name = 'UK&Ireland'
    else:
        # 其他情况按映射表查找
        dept_lower = dept.lower()
        for key, sheet in DEPT_SHEET_MAP.items():
            if key in dept_lower:
                sheet_name = sheet
                break

    if sheet_name not in wb.sheetnames:
        print(f"  [WARNING]  表3: sheet '{sheet_name}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    # 自动生成 Sr.No
    max_no = 0
    for r in range(2, next_row):
        v = ws.cell(r, 1).value
        if v and str(v).isdigit():
            max_no = max(max_no, int(v))
    data['sr_no'] = max_no + 1

    for field, col_letter in INV_COL_MAP.items():
        col_idx = column_index_from_string(col_letter)
        val = data.get(field, '')
        if not val:
            val = 'NA'
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        apply_style(cell, font_size=10)  # 表3: Calibri 10号

    wb.save(filepath)
    wb.close()
    print(f"  [OK] 表3 [{sheet_name}] 行 {next_row}: 写入成功 (Sr.No={data['sr_no']})")
    logger.info(f"    表3: 在 sheet '{sheet_name}' 第 {next_row} 行写入数据 (Sr.No={data['sr_no']})")
    return True


def main():
    # 执行前先备份原文件
    print("正在备份原文件...")
    backup_files()
    print()
    
    logger.info("=" * 65)
    logger.info("开始执行数据导入任务")
    logger.info(f"数据源: {PHOTOEXE_PATH}")
    logger.info(f"目标文件: {FAR_PATH}, {INV_PATH}")
    
    print("读取 photoexe.xlsx...")
    data_list = read_photoexe()
    print(f"共 {len(data_list)} 条记录\n")
    logger.info(f"成功读取 {len(data_list)} 条记录")

    for i, data in enumerate(data_list):
        print(f"--- 记录 #{i+1} ---")
        print(f"  SN: {data['serial_no']} | {data['brand']} {data['model']} | asset_id: {data['asset_id']} | global: {data['global_asset_tag']}")
        
        # 记录当前处理的数据
        logger.info(f"--- 处理记录 #{i+1} ---")
        logger.info(f"  城市: {data['city']}")
        logger.info(f"  部门: {data['department']}")
        logger.info(f"  设备类型: {data['type_of_equipment']}")
        logger.info(f"  品牌: {data['brand']}")
        logger.info(f"  型号: {data['model']}")
        logger.info(f"  序列号: {data['serial_no']}")
        logger.info(f"  Asset ID: {data['asset_id']}")
        logger.info(f"  Global Asset Tag: {data['global_asset_tag']}")

        print("  >> 写入表Asset_Far...")
        if write_to_far(data):
            logger.info(f"  成功写入表1 (Asset ID Status_FAR)")
        else:
            logger.error(f"  写入表1失败")

        print("  >> 写入表Inventory...")
        if write_to_inventory(data):
            logger.info(f"  成功写入表3 (VFS IT Inventory)")
        else:
            logger.error(f"  写入表3失败")

        print()
        logger.info("")  # 日志中加空行分隔

    print("=" * 65)
    print("完成!")
    logger.info("=" * 65)
    logger.info("数据导入任务完成")
    logger.info(f"共处理 {len(data_list)} 条记录")
    logger.info(f"日志已保存到: {LOG_FILE}")
    logger.info("=" * 65 + "\n")


if __name__ == "__main__":
    main()
