"""
VFS 资产管理 — 简化版
=========================
最简单的使用方式：

1. 准备数据（列表，每项是一个设备的字典）
2. 调用 write_assets(data) 即可

示例：
    data = [
        {
            'city': 'GZ',
            'office': 'VFS',
            'department': 'Canada',
            'type_of_equipment': 'Desktop',
            'brand': 'HP',
            'model': '280 G9',
            'serial_no': 'SN-001',
            'status_of_machine': 'In Use',
            'user_name': 'Tony'
        }
    ]
    write_assets(data)
"""

import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ============================================================
# 文件路径配置 - 修改这里指向你的文件
# ============================================================
FAR_PATH   = r"C:\Users\Administrator\Desktop\py\Asset ID Status_FAR.xlsx"
SPARE_PATH = r"C:\Users\Administrator\Desktop\py\China Spare Asset.xlsx"
INV_PATH   = r"C:\Users\Administrator\Desktop\py\VFS IT Inventory Guangzhou 2026.xlsx"

# ============================================================
# 列映射 - 不用管这里
# ============================================================
FAR_COL_MAP = {
    'city': 'A', 'office': 'B', 'department': 'C', 'type_of_equipment': 'D',
    'brand': 'E', 'model': 'F', 'serial_no': 'G', 'date_of_put_into_use': 'H',
    'cost_center': 'I', 'status_of_machine': 'J', 'asset_id_status': 'K',
    'asset_id': 'L', 'cc_id': 'M', 'entity': 'N', 'entity_name': 'O',
    'global_asset_tag': 'P', 'user_name': 'Q', 'hostname': 'R',
    'ip_address': 'S', 'warranty_amc': 'T', 'expiry_date': 'U',
    'processor': 'V', 'speed': 'W', 'ram': 'X', 'hdd': 'Y', 'size': 'Z',
    'os': 'AA', 'remark': 'AB'
}

SPARE_COL_MAP = {
    'sr_no': 'A', 'rm_name': 'B', 'owner': 'C', 'location': 'D',
    'city': 'E', 'office': 'F', 'department': 'G', 'type_of_equipment': 'H',
    'user_name': 'I', 'hostname': 'J', 'brand': 'K', 'model': 'L',
    'serial_no': 'M', 'date_of_put_into_use': 'N', 'city2': 'O',
    'office2': 'P', 'dept2': 'Q', 'status_of_machine': 'R',
    'recipient': 'S', 'remark': 'T'
}

INV_COL_MAP = {
    'sr_no': 'A', 'rm_name': 'B', 'fmc_or_not': 'C', 'location': 'D',
    'city': 'E', 'office': 'F', 'department': 'G', 'type_of_equipment': 'H',
    'user_name': 'I', 'hostname': 'J', 'ip_address': 'K', 'brand': 'L',
    'model': 'M', 'serial_no': 'N', 'asset_id': 'O', 'global_asset_tag': 'P',
    'purchase_date': 'Q', 'warranty_amc': 'R', 'expiry_date': 'S',
    'vendor': 'T', 'status_of_machine': 'U', 'processor': 'V', 'speed': 'W',
    'ram': 'X', 'hdd': 'Y', 'size': 'Z', 'os': 'AA', 'remark': 'AB',
    'purchase_value': 'AC', 'purchase_entity': 'AD', 'purchase_value_ref': 'AE'
}

# 部门 → sheet 映射
DEPT_SHEET_MAP = {
    'canada': 'Canada',
    'germany': 'Ger&JVAC&Swiss',
    'swiss': 'Ger&JVAC&Swiss',
    'jvac': 'Ger&JVAC&Swiss',
    'italy': 'Italy',
    'non-sch': 'Non-Sch',
    'platinum': 'Platinum Lounge',
    'ro': 'RO',
    'uk': 'UK&Ireland',
    'ireland': 'UK&Ireland',
    'us': 'US',
    'vasco': 'Vasco',
}


# ============================================================
# 核心函数 - 直接调用这些函数即可
# ============================================================

def write_to_far(data_dict, filepath=FAR_PATH, sheet_name=None):
    """写入表1 (Asset ID Status_FAR)"""
    if sheet_name is None:
        sheet_name = data_dict.get('city', 'GZ')

    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  sheet '{sheet_name}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    for field, col_letter in FAR_COL_MAP.items():
        if field in data_dict and data_dict[field]:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=data_dict[field])

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表1 [{sheet_name}] 行 {next_row}: 写入成功")
    return True


def write_to_spare(data_dict, filepath=SPARE_PATH):
    """写入表2 (China Spare Asset)"""
    wb = load_workbook(filepath)
    ws = wb['Raw Data']
    next_row = ws.max_row + 1

    # 自动生成 Sr.No
    if 'sr_no' not in data_dict or not data_dict['sr_no']:
        max_no = 0
        for r in range(2, next_row):
            v = ws.cell(r, 1).value
            if v and str(v).isdigit():
                max_no = max(max_no, int(v))
        data_dict['sr_no'] = max_no + 1

    for field, col_letter in SPARE_COL_MAP.items():
        if field in data_dict and data_dict[field]:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=data_dict[field])

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表2 [Raw Data] 行 {next_row}: 写入成功 (Sr.No={data_dict['sr_no']})")
    return True


def write_to_inventory(data_dict, filepath=INV_PATH, sheet_name=None):
    """写入表3 (VFS IT Inventory)"""
    wb = load_workbook(filepath)

    # 自动推断 sheet
    if sheet_name is None:
        dept = data_dict.get('department', '').lower()
        sheet_name = 'Canada'  # 默认
        for key, sheet in DEPT_SHEET_MAP.items():
            if key in dept:
                sheet_name = sheet
                break

    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  sheet '{sheet_name}' 不存在, 跳过")
        wb.close()
        return False

    ws = wb[sheet_name]
    next_row = ws.max_row + 1

    # 自动生成 Sr.No
    if 'sr_no' not in data_dict or not data_dict['sr_no']:
        max_no = 0
        for r in range(2, next_row):
            v = ws.cell(r, 1).value
            if v and str(v).isdigit():
                max_no = max(max_no, int(v))
        data_dict['sr_no'] = max_no + 1

    for field, col_letter in INV_COL_MAP.items():
        if field in data_dict and data_dict[field]:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx, value=data_dict[field])

    wb.save(filepath)
    wb.close()
    print(f"  ✅ 表3 [{sheet_name}] 行 {next_row}: 写入成功 (Sr.No={data_dict['sr_no']})")
    return True


def write_assets(data_list, target_tables=None):
    """
    主函数：写入资产数据到 Excel

    Args:
        data_list: 字典列表，每个字典是一项资产的数据
        target_tables: 要写入的表，可选 'far', 'spare', 'inv'，默认全部

    Returns:
        None

    示例：
        data = [
            {
                'city': 'GZ',
                'office': 'VFS',
                'department': 'Canada',
                'type_of_equipment': 'Desktop',
                'brand': 'HP',
                'model': '280 G9',
                'serial_no': 'SN-001',
                'status_of_machine': 'In Use',
                'user_name': 'Tony'
            }
        ]
        write_assets(data)
    """
    if target_tables is None:
        target_tables = ['far', 'spare', 'inv']

    print(f"\n{'='*65}")
    print(f"  写入 {len(data_list)} 条记录 → 目标: {target_tables}")
    print(f"{'='*65}\n")

    for i, data in enumerate(data_list):
        print(f"--- 记录 #{i+1} ---")
        print(f"  设备: {data.get('type_of_equipment', '')} | {data.get('brand', '')} {data.get('model', '')} | SN:{data.get('serial_no', '')}")

        if 'far' in target_tables:
            print("  >> 写入表1 (Asset ID Status_FAR)...")
            write_to_far(data)

        if 'spare' in target_tables:
            print("  >> 写入表2 (China Spare Asset)...")
            write_to_spare(data)

        if 'inv' in target_tables:
            print("  >> 写入表3 (VFS IT Inventory)...")
            write_to_inventory(data)

        print()

    print(f"{'='*65}")
    print(f"  完成!")
    print(f"{'='*65}\n")


# ============================================================
# 从 Excel 或 CSV 读取数据然后写入
# ============================================================
def write_from_excel(input_excel_path, target_tables=None):
    """
    从 Excel 文件读取数据并写入到三个表

    Args:
        input_excel_path: 输入 Excel 文件路径（第一行为表头）
        target_tables: 要写入的表，可选 'far', 'spare', 'inv'，默认全部
    """
    from openpyxl import load_workbook

    wb = load_workbook(input_excel_path)
    ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 读取数据
    data_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                data[headers[i]] = value
        if data:
            data_list.append(data)

    wb.close()

    # 写入
    write_assets(data_list, target_tables)


def write_from_csv(input_csv_path, target_tables=None):
    """
    从 CSV 文件读取数据并写入到三个表

    Args:
        input_csv_path: 输入 CSV 文件路径（第一行为表头）
        target_tables: 要写入的表，可选 'far', 'spare', 'inv'，默认全部
    """
    import csv

    data_list = []
    with open(input_csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 过滤空值
            data = {k: v for k, v in row.items() if v}
            if data:
                data_list.append(data)

    # 写入
    write_assets(data_list, target_tables)


# ============================================================
# 测试代码
# ============================================================
if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════════════╗
║       VFS 资产管理 — 简化版 (已就绪)                            ║
╚══════════════════════════════════════════════════════════════╝
    """)

    # 示例数据
    demo_data = [
        {
            'city': 'GZ',
            'office': 'VFS',
            'department': 'Canada',
            'type_of_equipment': 'Desktop',
            'brand': 'HP',
            'model': '280 G9 Demo',
            'serial_no': 'DEMO-001',
            'date_of_put_into_use': '2026-6',
            'status_of_machine': 'In Use',
            'user_name': 'Demo User 1',
            'processor': 'Intel Core i7',
            'ram': '16G',
            'os': 'WIN11',
            'remark': '演示数据'
        }
    ]

    # 写入演示数据
    write_assets(demo_data, target_tables=['far', 'spare', 'inv'])

    print("✅ 演示完成!")
    print("\n使用方式：")
    print("  1. 准备数据列表：")
    print("     data = [{'city': 'GZ', 'type_of_equipment': 'Desktop', ...}, ...]")
    print("  2. 调用写入函数：")
    print("     write_assets(data)")
    print("  3. 或者从 Excel/CSV 读取：")
    print("     write_from_excel('input.xlsx')")
    print("     write_from_csv('input.csv')")
